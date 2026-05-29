"""
storage/excel_exporter.py

Excel Export Layer

Responsibilities:
- export leads to xlsx
- append leads
- create sheets automatically
"""

from pathlib import Path

from openpyxl import (
    Workbook,
    load_workbook,
)

from models.lead_model import (
    LeadModel
)


class ExcelExporter:

    FILE_NAME = "data/leads.xlsx"

    HEADERS = [

        "Lead ID",

        "Company",

        "Website",

        "Emails",

        "Phones",

        "LinkedIn",

        "Address",

        "Source URL",
    ]

    @classmethod
    def _create_file(cls):

        path = Path(cls.FILE_NAME)

        if path.exists():
            return

        wb = Workbook()

        ws = wb.active

        ws.title = "Leads"

        ws.append(
            cls.HEADERS
        )

        wb.save(
            cls.FILE_NAME
        )

    @classmethod
    def export_lead(
        cls,
        lead: LeadModel,
    ):

        cls._create_file()

        wb = load_workbook(
            cls.FILE_NAME
        )

        ws = wb["Leads"]

        ws.append([

            lead.lead_id,

            lead.company.name,

            lead.company.website,

            ", ".join(
                lead.contact.emails
            ),

            ", ".join(
                lead.contact.phones
            ),

            ", ".join(
                lead.contact.linkedin_urls
            ),

            lead.address.full_address,

            lead.source_url,
        ])

        wb.save(
            cls.FILE_NAME
        )

    @classmethod
    def export_many(
        cls,
        leads: list[LeadModel],
    ):

        for lead in leads:

            cls.export_lead(
                lead
            )